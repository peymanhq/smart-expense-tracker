"""Application and atomic persistence tests for Excel import."""

from contextlib import contextmanager
from datetime import date, datetime, timezone
import json
from pathlib import Path

from openpyxl import Workbook, load_workbook
import pytest

from account import Account
from category import Category
from excel_exporter import export_transactions_to_excel
from excel_template import generate_excel_import_template
from excel_import_service import (
    ExcelImportPersistenceConflictError,
    ExcelImportPersistenceValidationError,
    ExcelImportService,
    InvalidExcelImportPreviewError,
)
from excel_workbook import REQUIRED_TRANSACTION_HEADERS
from json_storage import StorageError
import storage
from transaction import Transaction
from transaction_repository import JsonTransactionRepository
from transaction_service import TransactionService

TODAY = date(2026, 7, 27)
NOW = datetime(2026, 7, 27, 10, 30, tzinfo=timezone.utc)
ACCOUNT_ID = "123e4567-e89b-12d3-a456-426614174000"
OTHER_ACCOUNT_ID = "123e4567-e89b-12d3-a456-426614174010"
INCOME_CATEGORY_ID = "123e4567-e89b-12d3-a456-426614174001"
EXPENSE_CATEGORY_ID = "123e4567-e89b-12d3-a456-426614174002"
OTHER_EXPENSE_CATEGORY_ID = "123e4567-e89b-12d3-a456-426614174012"


def write_import(path: Path, rows) -> Path:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Transactions"
    worksheet.append(REQUIRED_TRANSACTION_HEADERS)
    for row in rows:
        worksheet.append(row)
    workbook.save(path)
    workbook.close()
    return path


def row(
    *,
    transaction_date="2026-07-20",
    transaction_type="Expense",
    amount=10,
    description="Lunch",
    account="Cash",
    category="Food",
):
    return (
        transaction_date,
        transaction_type,
        amount,
        description,
        account,
        category,
    )


@pytest.fixture
def accounts() -> list[Account]:
    return [
        Account(ACCOUNT_ID, "A-0001", "Cash"),
        Account(
            OTHER_ACCOUNT_ID,
            "A-0002",
            "Old Bank",
            is_active=False,
        ),
    ]


@pytest.fixture
def categories() -> list[Category]:
    return [
        Category(
            INCOME_CATEGORY_ID,
            "C-0001",
            "Salary",
            "income",
        ),
        Category(
            EXPENSE_CATEGORY_ID,
            "C-0002",
            "Food",
            "expense",
        ),
        Category(
            OTHER_EXPENSE_CATEGORY_ID,
            "C-0003",
            "Archived",
            "expense",
            is_active=False,
        ),
    ]


@pytest.fixture
def repository(tmp_path) -> JsonTransactionRepository:
    return JsonTransactionRepository(tmp_path / "data" / "transactions.json")


@pytest.fixture
def transaction_service(
    repository,
    accounts,
    categories,
) -> TransactionService:
    account_map = {account.id: account for account in accounts}
    category_map = {category.id: category for category in categories}
    return TransactionService(
        repository,
        today_provider=lambda: TODAY,
        utc_now_provider=lambda: NOW,
        account_lookup=account_map.get,
        category_lookup=category_map.get,
    )


@pytest.fixture
def import_service(
    transaction_service,
    accounts,
    categories,
) -> ExcelImportService:
    return ExcelImportService(
        transaction_service,
        account_list=lambda: list(accounts),
        category_list=lambda: list(categories),
    )


def test_active_names_resolve_case_insensitively_to_internal_uuids(
    tmp_path,
    import_service,
) -> None:
    source = write_import(
        tmp_path / "resolve.xlsx",
        [
            row(account=" cash ", category=" food "),
            row(
                transaction_type="Income",
                amount=100,
                description="Pay",
                category="SALARY",
            ),
        ],
    )

    preview = import_service.analyze(source)

    assert preview.is_valid
    assert [candidate.account_id for candidate in preview.candidates] == [
        ACCOUNT_ID,
        ACCOUNT_ID,
    ]
    assert [candidate.category_id for candidate in preview.candidates] == [
        EXPENSE_CATEGORY_ID,
        INCOME_CATEGORY_ID,
    ]
    assert preview.income_transaction_count == 1
    assert preview.expense_transaction_count == 1
    assert preview.total_income == 100
    assert preview.total_expense == 10
    assert preview.net_balance_impact == 90


def test_stale_template_category_after_type_change_is_rejected(
    tmp_path,
    import_service,
    accounts,
    categories,
) -> None:
    source = generate_excel_import_template(
        accounts,
        categories,
        tmp_path / "stale-category.xlsx",
    )
    workbook = load_workbook(source)
    worksheet = workbook["Transactions"]
    worksheet.append(
        row(
            transaction_type="Income",
            category="Food",
            description="Type changed after selecting Category",
        )
    )
    workbook.save(source)
    workbook.close()

    preview = import_service.analyze(source)

    assert not preview.is_valid
    assert preview.candidates == ()
    assert preview.issues[0].row_number == 2
    assert preview.issues[0].code == "category_type_mismatch"


@pytest.mark.parametrize(
    ("account", "category", "transaction_type", "code"),
    [
        ("Missing", "Food", "Expense", "unknown_account"),
        ("Old Bank", "Food", "Expense", "inactive_account"),
        ("Cash", "Missing", "Expense", "unknown_category"),
        ("Cash", "Archived", "Expense", "inactive_category"),
        ("Cash", "Salary", "Expense", "category_type_mismatch"),
        ("Cash", "Food", "Income", "category_type_mismatch"),
        (ACCOUNT_ID, "Food", "Expense", "unknown_account"),
        ("Cash", EXPENSE_CATEGORY_ID, "Expense", "unknown_category"),
    ],
)
def test_reference_errors_are_row_scoped_and_uuid_text_is_not_accepted(
    tmp_path,
    import_service,
    account,
    category,
    transaction_type,
    code,
) -> None:
    source = write_import(
        tmp_path / f"{code}.xlsx",
        [
            row(
                account=account,
                category=category,
                transaction_type=transaction_type,
            )
        ],
    )

    preview = import_service.analyze(source)

    assert not preview.is_valid
    assert preview.candidates == ()
    assert preview.issues[0].row_number == 2
    assert preview.issues[0].code == code
    assert str(preview.issues[0].supplied_value) in preview.issues[0].message


def test_future_date_is_reported_without_persistence(
    tmp_path,
    import_service,
    transaction_service,
) -> None:
    source = write_import(
        tmp_path / "future.xlsx",
        [row(transaction_date="2026-07-28")],
    )

    preview = import_service.analyze(source)

    assert preview.invalid_row_count == 1
    assert preview.issues[0].row_number == 2
    assert "cannot be after today" in preview.issues[0].message
    with pytest.raises(InvalidExcelImportPreviewError):
        import_service.persist(preview)
    assert transaction_service.list_transactions() == []


def test_duplicate_against_existing_reports_display_id(
    tmp_path,
    import_service,
    transaction_service,
) -> None:
    existing = transaction_service.add_transaction(
        transaction_date=date(2026, 7, 20),
        transaction_type="expense",
        amount=10,
        category="ignored",
        account="ignored",
        description="  LUNCH ",
        account_id=ACCOUNT_ID,
        category_id=EXPENSE_CATEGORY_ID,
    )
    source = write_import(tmp_path / "duplicate.xlsx", [row()])

    preview = import_service.analyze(source)

    assert preview.duplicate_conflict_count == 1
    assert preview.candidates == ()
    assert preview.issues[0].matching_display_id == existing.display_id
    assert existing.display_id in preview.issues[0].message


def test_duplicate_against_resolvable_legacy_name_snapshots_is_blocked(
    tmp_path,
    import_service,
    transaction_service,
) -> None:
    legacy = transaction_service.add_transaction(
        transaction_date=date(2026, 7, 20),
        transaction_type="expense",
        amount=10,
        category="Food",
        account="Cash",
        description="Lunch",
    )
    assert legacy.account_id is None
    assert legacy.category_id is None
    source = write_import(tmp_path / "legacy-duplicate.xlsx", [row()])

    preview = import_service.analyze(source)

    assert preview.duplicate_conflict_count == 1
    assert preview.issues[0].matching_display_id == legacy.display_id


def test_duplicate_inside_workbook_reports_earlier_excel_row(
    tmp_path,
    import_service,
) -> None:
    source = write_import(
        tmp_path / "batch-duplicate.xlsx",
        [
            row(description=" Lunch "),
            row(description="LUNCH"),
        ],
    )

    preview = import_service.analyze(source)

    assert preview.duplicate_conflict_count == 1
    assert preview.valid_candidate_count == 1
    assert preview.issues[0].row_number == 3
    assert preview.issues[0].earlier_row_number == 2


@pytest.mark.parametrize(
    "change",
    [
        {"transaction_date": "2026-07-21"},
        {"transaction_type": "Income", "category": "Salary"},
        {"amount": 10.01},
        {"description": "Dinner"},
        {"account": "Second"},
        {"category": "Travel"},
    ],
)
def test_difference_in_any_comparison_key_field_is_not_duplicate(
    tmp_path,
    repository,
    accounts,
    categories,
    change,
) -> None:
    second_account = Account(
        OTHER_ACCOUNT_ID,
        "A-0002",
        "Second",
    )
    travel = Category(
        OTHER_EXPENSE_CATEGORY_ID,
        "C-0003",
        "Travel",
        "expense",
    )
    all_accounts = [accounts[0], second_account]
    all_categories = [*categories[:2], travel]
    account_map = {item.id: item for item in all_accounts}
    category_map = {item.id: item for item in all_categories}
    service = TransactionService(
        repository,
        today_provider=lambda: TODAY,
        utc_now_provider=lambda: NOW,
        account_lookup=account_map.get,
        category_lookup=category_map.get,
    )
    service.add_transaction(
        transaction_date=date(2026, 7, 20),
        transaction_type="expense",
        amount=10,
        category="Food",
        account="Cash",
        description="Lunch",
        account_id=ACCOUNT_ID,
        category_id=EXPENSE_CATEGORY_ID,
    )
    importer = ExcelImportService(
        service,
        account_list=lambda: all_accounts,
        category_list=lambda: all_categories,
    )
    values = {
        "transaction_date": "2026-07-20",
        "transaction_type": "Expense",
        "amount": 10,
        "description": "Lunch",
        "account": "Cash",
        "category": "Food",
        **change,
    }
    source = write_import(tmp_path / "different.xlsx", [row(**values)])

    preview = importer.analyze(source)

    assert preview.is_valid
    assert preview.valid_candidate_count == 1


def test_resolved_account_and_category_uuids_are_part_of_duplicate_key(
    tmp_path,
    repository,
) -> None:
    first_account = Account(ACCOUNT_ID, "A-0001", "Cash")
    second_account = Account(OTHER_ACCOUNT_ID, "A-0002", "Cash")
    first_category = Category(
        EXPENSE_CATEGORY_ID,
        "C-0001",
        "Food",
        "expense",
    )
    second_category = Category(
        OTHER_EXPENSE_CATEGORY_ID,
        "C-0002",
        "Food",
        "expense",
    )
    lookup_accounts = {
        first_account.id: first_account,
        second_account.id: second_account,
    }
    lookup_categories = {
        first_category.id: first_category,
        second_category.id: second_category,
    }
    service = TransactionService(
        repository,
        today_provider=lambda: TODAY,
        utc_now_provider=lambda: NOW,
        account_lookup=lookup_accounts.get,
        category_lookup=lookup_categories.get,
    )
    service.add_transaction(
        transaction_date=date(2026, 7, 20),
        transaction_type="expense",
        amount=10,
        category="Food",
        account="Cash",
        description="Lunch",
        account_id=first_account.id,
        category_id=first_category.id,
    )
    importer = ExcelImportService(
        service,
        account_list=lambda: [second_account],
        category_list=lambda: [second_category],
    )
    source = write_import(tmp_path / "different-ids.xlsx", [row()])

    preview = importer.analyze(source)

    assert preview.is_valid


def test_atomic_import_generates_new_identity_timestamps_and_ordered_ids(
    tmp_path,
    import_service,
    transaction_service,
    monkeypatch,
) -> None:
    generated_ids = iter(
        [
            "123e4567-e89b-12d3-a456-426614174101",
            "123e4567-e89b-12d3-a456-426614174102",
        ]
    )
    monkeypatch.setattr(
        "transaction_factory.generator_transaction_id",
        lambda: next(generated_ids),
    )
    source = write_import(
        tmp_path / "identity.xlsx",
        [
            row(description="First"),
            row(amount=20, description="Second"),
        ],
    )

    result = import_service.persist(import_service.analyze(source))

    assert result.imported_count == 2
    assert [item.display_id for item in result.transactions] == [
        "T-0001",
        "T-0002",
    ]
    assert [item.id for item in result.transactions] == [
        "123e4567-e89b-12d3-a456-426614174101",
        "123e4567-e89b-12d3-a456-426614174102",
    ]
    assert [item.description for item in result.transactions] == [
        "First",
        "Second",
    ]
    assert all(item.created_at == NOW for item in result.transactions)
    assert all(item.updated_at == NOW for item in result.transactions)
    assert transaction_service.list_transactions() == list(result.transactions)


def test_deleted_display_ids_are_not_reused_by_bulk_import(
    tmp_path,
    import_service,
    transaction_service,
) -> None:
    created = [
        transaction_service.add_transaction(
            transaction_date=date(2026, 7, 20),
            transaction_type="expense",
            amount=index,
            category="Food",
            account="Cash",
            description=f"Existing {index}",
            account_id=ACCOUNT_ID,
            category_id=EXPENSE_CATEGORY_ID,
        )
        for index in (1, 2, 3)
    ]
    transaction_service.delete_transaction(
        created[-1].display_id,
        active_date=date(2026, 7, 20),
    )
    source = write_import(
        tmp_path / "after-delete.xlsx",
        [
            row(amount=4, description="Fourth"),
            row(amount=5, description="Fifth"),
        ],
    )

    result = import_service.persist(import_service.analyze(source))

    assert [item.display_id for item in result.transactions] == [
        "T-0004",
        "T-0005",
    ]


def test_bulk_import_uses_one_transaction_lock(
    tmp_path,
    import_service,
    monkeypatch,
) -> None:
    source = write_import(
        tmp_path / "one-lock.xlsx",
        [
            row(description="First"),
            row(amount=20, description="Second"),
        ],
    )
    preview = import_service.analyze(source)
    original_lock = storage.transaction_file_lock
    calls = 0

    @contextmanager
    def counting_lock(data_file=None):
        nonlocal calls
        calls += 1
        with original_lock(data_file):
            yield

    monkeypatch.setattr(storage, "transaction_file_lock", counting_lock)

    import_service.persist(preview)

    assert calls == 1


def test_persistence_failure_preserves_original_document(
    tmp_path,
    import_service,
    transaction_service,
    repository,
    monkeypatch,
) -> None:
    transaction_service.add_transaction(
        transaction_date=date(2026, 7, 20),
        transaction_type="expense",
        amount=5,
        category="Food",
        account="Cash",
        description="Existing",
        account_id=ACCOUNT_ID,
        category_id=EXPENSE_CATEGORY_ID,
    )
    data_file = repository._data_file
    original = data_file.read_text(encoding="utf-8")
    source = write_import(
        tmp_path / "failure.xlsx",
        [row(description="New")],
    )
    preview = import_service.analyze(source)

    def fail_write(document, data_file=None):
        raise StorageError("disk full")

    monkeypatch.setattr(storage, "_write_document", fail_write)

    with pytest.raises(StorageError, match="disk full"):
        import_service.persist(preview)
    assert data_file.read_text(encoding="utf-8") == original


def test_late_duplicate_after_preview_blocks_entire_import(
    tmp_path,
    import_service,
    transaction_service,
) -> None:
    source = write_import(
        tmp_path / "late-conflict.xlsx",
        [
            row(description="First"),
            row(amount=20, description="Second"),
        ],
    )
    preview = import_service.analyze(source)
    transaction_service.add_transaction(
        transaction_date=date(2026, 7, 20),
        transaction_type="expense",
        amount=10,
        category="Food",
        account="Cash",
        description="First",
        account_id=ACCOUNT_ID,
        category_id=EXPENSE_CATEGORY_ID,
    )
    before = transaction_service.list_transactions()

    with pytest.raises(
        ExcelImportPersistenceConflictError,
        match="existing transaction T-0001",
    ):
        import_service.persist(preview)

    assert transaction_service.list_transactions() == before


def test_reference_deactivated_after_preview_reports_row_and_imports_nothing(
    tmp_path,
    import_service,
    transaction_service,
    accounts,
) -> None:
    source = write_import(
        tmp_path / "late-inactive.xlsx",
        [row(description="First")],
    )
    preview = import_service.analyze(source)
    accounts[0].is_active = False

    with pytest.raises(
        ExcelImportPersistenceValidationError,
        match="Row 2: data changed after preview.*inactive",
    ):
        import_service.persist(preview)

    assert transaction_service.list_transactions() == []


def test_importing_same_workbook_twice_is_blocked(
    tmp_path,
    import_service,
    transaction_service,
) -> None:
    source = write_import(
        tmp_path / "twice.xlsx",
        [row(description="Only")],
    )
    first_preview = import_service.analyze(source)
    import_service.persist(first_preview)

    second_preview = import_service.analyze(source)

    assert second_preview.duplicate_conflict_count == 1
    with pytest.raises(InvalidExcelImportPreviewError):
        import_service.persist(second_preview)
    assert len(transaction_service.list_transactions()) == 1


def test_candidate_document_advances_metadata_once(tmp_path, import_service) -> None:
    source = write_import(
        tmp_path / "metadata.xlsx",
        [
            row(description="First"),
            row(amount=20, description="Second"),
        ],
    )
    result = import_service.persist(import_service.analyze(source))
    data_file = import_service._transaction_service._repository._data_file
    document = json.loads(data_file.read_text(encoding="utf-8"))

    assert len(result.transactions) == 2
    assert document["metadata"]["next_display_id"] == 3
    assert [item["display_id"] for item in document["transactions"]] == [
        "T-0001",
        "T-0002",
    ]


def test_current_export_is_import_compatible_and_identity_metadata_is_ignored(
    tmp_path,
    import_service,
) -> None:
    old_created = datetime(2000, 1, 1, tzinfo=timezone.utc)
    exported = Transaction(
        id="untrusted-internal-id",
        display_id="T-9999",
        type="expense",
        amount=15,
        category="Food",
        account="Cash",
        description="Exported",
        transaction_date=date(2026, 7, 20),
        created_at=old_created,
        updated_at=old_created,
        account_id=EXPENSE_CATEGORY_ID,
        category_id=ACCOUNT_ID,
    )
    source = export_transactions_to_excel(
        [exported],
        tmp_path / "export.xlsx",
        account_names={EXPENSE_CATEGORY_ID: "Cash"},
        category_names={ACCOUNT_ID: "Food"},
    )

    result = import_service.persist(import_service.analyze(source))

    assert result.transactions[0].id != exported.id
    assert result.transactions[0].display_id == "T-0001"
    assert result.transactions[0].created_at == NOW
    assert result.transactions[0].updated_at == NOW
