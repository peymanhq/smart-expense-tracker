"""Workbook-level tests for the Excel import template."""

from pathlib import Path

from openpyxl import load_workbook
import pytest

from account import Account
from category import Category
from excel_template import (
    INSTRUCTIONS,
    TEMPLATE_WORKSHEETS,
    generate_excel_import_template,
)
from excel_workbook import (
    AMOUNT_FORMAT,
    DATE_FORMAT,
    REQUIRED_TRANSACTION_HEADERS,
    ExcelDestinationExistsError,
    ExcelSaveError,
    InvalidExcelDestinationError,
)
import excel_workbook

ACCOUNT_ID = "123e4567-e89b-12d3-a456-426614174000"
INACTIVE_ACCOUNT_ID = "123e4567-e89b-12d3-a456-426614174010"
INCOME_CATEGORY_ID = "123e4567-e89b-12d3-a456-426614174001"
EXPENSE_CATEGORY_ID = "123e4567-e89b-12d3-a456-426614174002"
INACTIVE_CATEGORY_ID = "123e4567-e89b-12d3-a456-426614174012"


def records():
    return (
        [
            Account(ACCOUNT_ID, "A-0001", "Daily Cash"),
            Account(
                INACTIVE_ACCOUNT_ID,
                "A-0002",
                "Closed / Archive",
                is_active=False,
            ),
        ],
        [
            Category(
                INCOME_CATEGORY_ID,
                "C-0001",
                "Salary & Bonus",
                "income",
            ),
            Category(
                EXPENSE_CATEGORY_ID,
                "C-0002",
                "Food / Dining",
                "expense",
            ),
            Category(
                INACTIVE_CATEGORY_ID,
                "C-0003",
                "Inactive",
                "expense",
                is_active=False,
            ),
        ],
    )


def workbook_values(workbook) -> list[object]:
    return [
        cell.value
        for worksheet in workbook.worksheets
        for row in worksheet.iter_rows()
        for cell in row
        if cell.value is not None
    ]


def test_template_creates_expected_professional_workbook(tmp_path) -> None:
    accounts, categories = records()
    destination = tmp_path / "template.xlsx"

    result = generate_excel_import_template(
        accounts,
        categories,
        destination,
    )

    assert result == destination
    workbook = load_workbook(destination)
    assert tuple(workbook.sheetnames) == TEMPLATE_WORKSHEETS

    transactions = workbook["Transactions"]
    assert tuple(cell.value for cell in transactions[1]) == (
        REQUIRED_TRANSACTION_HEADERS
    )
    assert transactions.max_row == 1
    assert transactions.freeze_panes == "A2"
    assert transactions.auto_filter.ref == "A1:F1"
    assert transactions.column_dimensions["A"].number_format == DATE_FORMAT
    assert transactions.column_dimensions["C"].number_format == AMOUNT_FORMAT
    assert transactions["A1"].fill.fgColor.rgb.endswith("1F4E78")
    assert transactions["A1"].font.bold is True

    validations = {
        validation.formula1: str(validation.sqref)
        for validation in transactions.data_validations.dataValidation
    }
    assert validations['"Income,Expense"'] == "B2:B1001"
    assert validations["=ActiveAccounts"] == "E2:E1001"
    assert validations["=ActiveCategories"] == "F2:F1001"
    workbook.close()


def test_instructions_include_every_required_warning(tmp_path) -> None:
    accounts, categories = records()
    destination = generate_excel_import_template(
        accounts,
        categories,
        tmp_path / "instructions.xlsx",
    )
    workbook = load_workbook(destination, read_only=True)
    text = " ".join(
        str(value)
        for row in workbook["Instructions"].iter_rows(values_only=True)
        for value in row
        if value is not None
    ).casefold()

    required_phrases = (
        "transactions",
        "date, type, amount, description, account, and category",
        "income or expense",
        "yyyy-mm-dd",
        "greater than zero",
        "active account",
        "active category",
        "match the transaction type",
        "completely empty rows",
        "duplicate",
        "all-or-nothing",
        "do not rename",
        "do not use a uuid",
        "display ids",
        "created at",
        "updated at",
        ".xlsx",
    )
    assert len(INSTRUCTIONS) >= 12
    assert all(phrase in text for phrase in required_phrases)
    workbook.close()


def test_reference_data_contains_only_active_names_and_no_internal_ids(
    tmp_path,
) -> None:
    accounts, categories = records()
    destination = generate_excel_import_template(
        accounts,
        categories,
        tmp_path / "references.xlsx",
    )
    workbook = load_workbook(destination)
    reference = workbook["Reference Data"]

    assert reference["A1"].value == "Active Accounts"
    assert reference["A2"].value == "Daily Cash"
    assert reference["A3"].value is None
    assert reference["C1"].value == "Active Income Categories"
    assert reference["C2"].value == "Salary & Bonus"
    assert reference["E1"].value == "Active Expense Categories"
    assert reference["E2"].value == "Food / Dining"
    assert reference["G2"].value == "Salary & Bonus"
    assert reference["G3"].value == "Food / Dining"
    values = workbook_values(workbook)
    assert "Closed / Archive" not in values
    assert "Inactive" not in values
    assert ACCOUNT_ID not in values
    assert INCOME_CATEGORY_ID not in values
    assert EXPENSE_CATEGORY_ID not in values

    names = {name.name: name.attr_text for name in workbook.defined_names.values()}
    assert names["ActiveAccounts"] == "'Reference Data'!$A$2:$A$2"
    assert names["ActiveIncomeCategories"] == "'Reference Data'!$C$2:$C$2"
    assert names["ActiveExpenseCategories"] == "'Reference Data'!$E$2:$E$2"
    assert names["ActiveCategories"] == "'Reference Data'!$G$2:$G$3"
    workbook.close()


def test_empty_reference_lists_have_valid_blank_named_ranges(tmp_path) -> None:
    destination = generate_excel_import_template(
        [],
        [],
        tmp_path / "empty.xlsx",
    )
    workbook = load_workbook(destination)
    names = {name.name: name.attr_text for name in workbook.defined_names.values()}

    assert names["ActiveAccounts"] == "'Reference Data'!$A$2:$A$2"
    assert names["ActiveCategories"] == "'Reference Data'!$G$2:$G$2"
    assert workbook["Reference Data"]["A2"].value is None
    assert workbook["Reference Data"]["G2"].value is None
    workbook.close()


def test_destination_normalization_overwrite_and_extension_policy(
    tmp_path,
) -> None:
    result = generate_excel_import_template([], [], tmp_path / "nested" / "name")
    assert result == tmp_path / "nested" / "name.xlsx"

    original = result.read_bytes()
    with pytest.raises(ExcelDestinationExistsError):
        generate_excel_import_template([], [], result)
    assert result.read_bytes() == original

    generate_excel_import_template([], [], result, overwrite=True)
    assert load_workbook(result).sheetnames[0] == "Instructions"

    with pytest.raises(InvalidExcelDestinationError):
        generate_excel_import_template([], [], tmp_path / "wrong.xls")
    with pytest.raises(InvalidExcelDestinationError):
        generate_excel_import_template([], [], tmp_path / "wrong.xlsx.csv")


def test_save_failure_leaves_no_partial_output(
    tmp_path,
    monkeypatch,
) -> None:
    destination = tmp_path / "failed.xlsx"

    def fail_save(self, filename):
        raise OSError("disk unavailable")

    monkeypatch.setattr(excel_workbook.Workbook, "save", fail_save)

    with pytest.raises(ExcelSaveError, match="disk unavailable"):
        generate_excel_import_template([], [], destination)

    assert not destination.exists()
    assert list(tmp_path.iterdir()) == []


def test_failed_overwrite_preserves_existing_template(
    tmp_path,
    monkeypatch,
) -> None:
    destination = tmp_path / "existing.xlsx"
    destination.write_bytes(b"original")

    def fail_save(self, filename):
        raise OSError("disk unavailable")

    monkeypatch.setattr(excel_workbook.Workbook, "save", fail_save)
    with pytest.raises(ExcelSaveError):
        generate_excel_import_template([], [], destination, overwrite=True)

    assert destination.read_bytes() == b"original"
    assert list(tmp_path.iterdir()) == [destination]
