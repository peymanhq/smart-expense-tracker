"""Workbook parsing tests for Excel transaction import."""

from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook
import pytest

from excel_import import (
    ExcelImportFileNotFoundError,
    InvalidExcelHeadersError,
    InvalidExcelWorkbookError,
    MissingTransactionsWorksheetError,
    UnsupportedExcelImportFileError,
    parse_excel_transactions,
)
import excel_import
from excel_workbook import REQUIRED_TRANSACTION_HEADERS


def write_workbook(
    path: Path,
    rows=(),
    *,
    headers=REQUIRED_TRANSACTION_HEADERS,
    sheet_name="Transactions",
) -> Path:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_name
    worksheet.append(headers)
    for row in rows:
        worksheet.append(row)
    workbook.save(path)
    workbook.close()
    return path


def valid_row(**overrides):
    values = {
        "Date": "2026-07-20",
        "Type": "Expense",
        "Amount": 12.5,
        "Description": " Lunch ",
        "Account": " Cash ",
        "Category": " Food ",
    }
    values.update(overrides)
    return tuple(values[header] for header in REQUIRED_TRANSACTION_HEADERS)


def issue_messages(result) -> list[str]:
    return [issue.message for issue in result.issues]


def test_valid_workbook_parses_and_normalizes_fields(tmp_path) -> None:
    source = write_workbook(
        tmp_path / "valid.xlsx",
        [
            valid_row(
                Date=date(2026, 7, 19),
                Type=" income ",
                Amount="125.50",
                Description=" Pay ",
                Account=" Bank ",
                Category=" Salary ",
            ),
            valid_row(
                Date=datetime(2026, 7, 20, 8, 30),
                Type="EXPENSE",
                Amount=12,
            ),
        ],
    )

    result = parse_excel_transactions(source)

    assert result.source_path == source
    assert result.worksheet_name == "Transactions"
    assert result.total_physical_data_rows == 2
    assert result.empty_rows_ignored == 0
    assert result.issues == ()
    assert [row.row_number for row in result.rows] == [2, 3]
    assert [row.transaction_date for row in result.rows] == [
        date(2026, 7, 19),
        date(2026, 7, 20),
    ]
    assert result.rows[0].transaction_type == "income"
    assert result.rows[0].amount == pytest.approx(125.5)
    assert result.rows[0].description == "Pay"
    assert result.rows[0].account_name == "Bank"
    assert result.rows[0].category_name == "Salary"


@pytest.mark.parametrize("name", ["input.xls", "input.xlsm", "input.xlsx.csv"])
def test_non_xlsx_and_misleading_extensions_are_rejected(
    tmp_path,
    name,
) -> None:
    source = tmp_path / name
    source.write_bytes(b"not relevant")

    with pytest.raises(UnsupportedExcelImportFileError, match=".xlsx"):
        parse_excel_transactions(source)


def test_missing_and_corrupt_workbooks_raise_controlled_errors(tmp_path) -> None:
    with pytest.raises(ExcelImportFileNotFoundError):
        parse_excel_transactions(tmp_path / "missing.xlsx")

    corrupt = tmp_path / "corrupt.xlsx"
    corrupt.write_bytes(b"not an Excel archive")
    with pytest.raises(InvalidExcelWorkbookError) as captured:
        parse_excel_transactions(corrupt)
    assert captured.value.__cause__ is not None


def test_missing_transactions_worksheet_is_rejected(tmp_path) -> None:
    source = write_workbook(
        tmp_path / "wrong-sheet.xlsx",
        sheet_name="transactions",
    )

    with pytest.raises(MissingTransactionsWorksheetError, match="exactly"):
        parse_excel_transactions(source)


def test_headers_are_trimmed_case_insensitive_and_allow_extras(tmp_path) -> None:
    headers = (
        " extra ",
        " date ",
        "TYPE",
        "Amount",
        "Description",
        "ACCOUNT",
        " category ",
        "Display ID",
        "Created At",
    )
    source = write_workbook(
        tmp_path / "headers.xlsx",
        [
            (
                "ignored",
                "2026-07-20",
                "Expense",
                5,
                "Coffee",
                "Cash",
                "Food",
                "T-9999",
                "1900-01-01",
            )
        ],
        headers=headers,
    )

    result = parse_excel_transactions(source)

    assert result.issues == ()
    assert len(result.rows) == 1
    assert result.rows[0].amount == 5


def test_legacy_export_transaction_date_header_is_supported(tmp_path) -> None:
    headers = (
        "Display ID",
        "Transaction Date",
        "Type",
        "Amount",
        "Description",
        "Account",
        "Category",
        "Created At",
        "Updated At",
    )
    source = write_workbook(
        tmp_path / "legacy-export.xlsx",
        [
            (
                "T-0001",
                "2026-07-20",
                "Expense",
                5,
                "Coffee",
                "Cash",
                "Food",
                "ignored",
                "ignored",
            )
        ],
        headers=headers,
    )

    result = parse_excel_transactions(source)

    assert result.issues == ()
    assert result.rows[0].transaction_date == date(2026, 7, 20)


def test_missing_and_duplicate_normalized_headers_are_rejected(tmp_path) -> None:
    missing = write_workbook(
        tmp_path / "missing-header.xlsx",
        headers=REQUIRED_TRANSACTION_HEADERS[:-1],
    )
    with pytest.raises(InvalidExcelHeadersError, match="Category"):
        parse_excel_transactions(missing)

    duplicate = write_workbook(
        tmp_path / "duplicate-header.xlsx",
        headers=(*REQUIRED_TRANSACTION_HEADERS, " amount "),
    )
    with pytest.raises(InvalidExcelHeadersError, match="duplicate"):
        parse_excel_transactions(duplicate)

    ambiguous_date = write_workbook(
        tmp_path / "ambiguous-date.xlsx",
        headers=(*REQUIRED_TRANSACTION_HEADERS, "Transaction Date"),
    )
    with pytest.raises(InvalidExcelHeadersError, match="ambiguous"):
        parse_excel_transactions(ambiguous_date)


def test_completely_empty_rows_are_ignored_and_row_numbers_preserved(
    tmp_path,
) -> None:
    source = write_workbook(
        tmp_path / "empty-row.xlsx",
        [
            valid_row(),
            (None, None, None, None, None, None),
            valid_row(Amount=20),
        ],
    )

    result = parse_excel_transactions(source)

    assert result.total_physical_data_rows == 3
    assert result.empty_rows_ignored == 1
    assert [row.row_number for row in result.rows] == [2, 4]


def test_partial_rows_report_every_invalid_field_with_excel_row(tmp_path) -> None:
    source = write_workbook(
        tmp_path / "partial.xlsx",
        [
            (
                None,
                "",
                0,
                "description",
                " ",
                None,
            )
        ],
    )

    result = parse_excel_transactions(source)

    assert result.rows == ()
    assert {issue.row_number for issue in result.issues} == {2}
    messages = issue_messages(result)
    assert any("Date is required" in message for message in messages)
    assert "Type is required." in messages
    assert "Amount must be a valid number greater than zero." in messages
    assert "Account is required." in messages
    assert "Category is required." in messages


@pytest.mark.parametrize(
    ("field", "value", "message"),
    [
        ("Date", "2026-02-30", "Date must be a valid"),
        ("Date", "20/07/2026", "Date must be a valid"),
        ("Date", True, "Date must be a real"),
        ("Type", "Transfer", "Income or Expense"),
        ("Type", True, "Income or Expense"),
        ("Amount", -1, "greater than zero"),
        ("Amount", True, "finite number"),
        ("Amount", "nan", "finite number"),
        ("Amount", "inf", "finite number"),
        ("Amount", "not numeric", "valid number"),
        ("Description", 42, "Description must be text"),
    ],
)
def test_field_validation_reports_invalid_values(
    tmp_path,
    field,
    value,
    message,
) -> None:
    source = write_workbook(
        tmp_path / f"invalid-{field}.xlsx",
        [valid_row(**{field: value})],
    )

    result = parse_excel_transactions(source)

    assert result.rows == ()
    assert result.issues[0].row_number == 2
    assert message in result.issues[0].message


@pytest.mark.parametrize("field", ["Date", "Amount"])
def test_formula_without_cached_value_is_rejected(tmp_path, field) -> None:
    source = tmp_path / f"formula-{field}.xlsx"
    write_workbook(source, [valid_row()])
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Transactions"
    worksheet.append(REQUIRED_TRANSACTION_HEADERS)
    worksheet.append(valid_row())
    column = REQUIRED_TRANSACTION_HEADERS.index(field) + 1
    worksheet.cell(row=2, column=column, value="=1+1")
    workbook.save(source)
    workbook.close()

    result = parse_excel_transactions(source)

    assert result.rows == ()
    assert "cached value" in result.issues[0].message


def test_blank_description_uses_existing_optional_text_semantics(tmp_path) -> None:
    source = write_workbook(
        tmp_path / "blank-description.xlsx",
        [valid_row(Description=None)],
    )

    result = parse_excel_transactions(source)

    assert result.issues == ()
    assert result.rows[0].description == ""


@pytest.mark.parametrize(
    ("sheet_name", "raises"),
    [
        ("Transactions", None),
        ("Wrong", MissingTransactionsWorksheetError),
    ],
)
def test_workbook_resource_closes_on_success_and_validation_failure(
    tmp_path,
    monkeypatch,
    sheet_name,
    raises,
) -> None:
    source = write_workbook(
        tmp_path / f"close-{sheet_name}.xlsx",
        [valid_row()],
        sheet_name=sheet_name,
    )
    original_loader = excel_import.load_workbook
    closed = []

    def tracking_loader(*args, **kwargs):
        workbook = original_loader(*args, **kwargs)
        original_close = workbook.close

        def close():
            closed.append(True)
            original_close()

        workbook.close = close
        return workbook

    monkeypatch.setattr(excel_import, "load_workbook", tracking_loader)

    if raises is None:
        parse_excel_transactions(source)
    else:
        with pytest.raises(raises):
            parse_excel_transactions(source)
    assert closed == [True]


def test_loader_uses_defensive_read_options(tmp_path, monkeypatch) -> None:
    source = write_workbook(
        tmp_path / "options.xlsx",
        [valid_row()],
    )
    original_loader = excel_import.load_workbook
    captured = {}

    def recording_loader(*args, **kwargs):
        captured.update(kwargs)
        return original_loader(*args, **kwargs)

    monkeypatch.setattr(excel_import, "load_workbook", recording_loader)

    parse_excel_transactions(source)

    assert captured == {
        "read_only": True,
        "data_only": True,
        "keep_links": False,
    }
