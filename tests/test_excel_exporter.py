"""Workbook-level tests for the Excel reporting adapter."""

from copy import deepcopy
from datetime import date, datetime, timezone
from pathlib import Path

from openpyxl import load_workbook
import pytest

import excel_exporter
from excel_exporter import (
    CATEGORY_SUMMARY_HEADERS,
    SUMMARY_HEADERS,
    TRANSACTION_HEADERS,
    ExcelDestinationExistsError,
    ExcelSaveError,
    InvalidExcelDestinationError,
    export_transactions_to_excel,
)
from transaction import Transaction


def make_transaction(
    number: int,
    transaction_type: str,
    amount: float,
    *,
    transaction_date: date = date(2026, 7, 20),
    account: str = "Cash snapshot",
    category: str = "General snapshot",
    account_id: str | None = None,
    category_id: str | None = None,
    created_at: datetime | None = None,
    updated_at: datetime | None = None,
) -> Transaction:
    return Transaction(
        id=f"internal-{number}",
        display_id=f"T-{number:04d}",
        type=transaction_type,
        amount=amount,
        category=category,
        account=account,
        description=f"Description {number}",
        transaction_date=transaction_date,
        created_at=created_at,
        updated_at=updated_at,
        account_id=account_id,
        category_id=category_id,
    )


def open_export(tmp_path: Path, transactions, **kwargs):
    destination = tmp_path / "report.xlsx"
    result = export_transactions_to_excel(
        transactions,
        destination,
        **kwargs,
    )
    return result, load_workbook(destination)


def test_export_creates_reopenable_workbook_with_deterministic_sheets(
    tmp_path,
) -> None:
    result, workbook = open_export(tmp_path, [])

    assert result == tmp_path / "report.xlsx"
    assert result.is_file()
    assert workbook.sheetnames == [
        "Transactions",
        "Summary",
        "Category Summary",
    ]
    assert tuple(
        cell.value for cell in workbook["Transactions"][1]
    ) == TRANSACTION_HEADERS
    assert tuple(cell.value for cell in workbook["Summary"][1]) == SUMMARY_HEADERS
    assert tuple(
        cell.value for cell in workbook["Category Summary"][1]
    ) == CATEGORY_SUMMARY_HEADERS
    workbook.close()


def test_transaction_rows_are_complete_resolved_and_deterministic(
    tmp_path,
) -> None:
    created = datetime(2026, 7, 21, 10, 30, tzinfo=timezone.utc)
    updated = datetime(2026, 7, 21, 11, 45, tzinfo=timezone.utc)
    transactions = [
        make_transaction(
            2,
            "expense",
            19.99,
            transaction_date=date(2026, 7, 21),
            account_id="account-2",
            category_id="category-2",
            created_at=created,
            updated_at=updated,
        ),
        make_transaction(
            1,
            "income",
            1234.56,
            transaction_date=date(2026, 7, 20),
            account="Legacy Cash",
            category="Legacy Salary",
        ),
    ]
    _, workbook = open_export(
        tmp_path,
        transactions,
        account_names={"account-2": "Current Account"},
        category_names={"category-2": "Dining"},
    )
    worksheet = workbook["Transactions"]

    assert worksheet.max_row == 3
    assert [worksheet.cell(row, 1).value for row in (2, 3)] == [
        "T-0001",
        "T-0002",
    ]
    assert worksheet["B2"].value.date() == date(2026, 7, 20)
    assert worksheet["B2"].value.date() != created.date()
    assert worksheet["C3"].value == "Expense"
    assert worksheet["D2"].value == pytest.approx(1234.56)
    assert worksheet["D3"].value == pytest.approx(19.99)
    assert worksheet["E3"].value == "Description 2"
    assert worksheet["F2"].value == "Legacy Cash"
    assert worksheet["G2"].value == "Legacy Salary"
    assert worksheet["F3"].value == "Current Account"
    assert worksheet["G3"].value == "Dining"
    assert worksheet["H3"].value == created.replace(tzinfo=None)
    assert worksheet["I3"].value == updated.replace(tzinfo=None)
    assert worksheet["H2"].value is None
    assert worksheet["B3"].number_format == "yyyy-mm-dd"
    assert worksheet["D3"].number_format == "#,##0.00"
    assert worksheet["H3"].number_format == "yyyy-mm-dd hh:mm:ss"
    workbook.close()


@pytest.mark.parametrize("amount", [0.10, 19.99, 1234.56])
def test_export_preserves_existing_float_amount_semantics(
    tmp_path,
    amount,
) -> None:
    _, workbook = open_export(
        tmp_path,
        [make_transaction(1, "expense", amount)],
    )

    assert workbook["Transactions"]["D2"].value == pytest.approx(amount)
    workbook.close()


def test_unresolved_managed_names_are_explicitly_blank(tmp_path) -> None:
    transaction = make_transaction(
        1,
        "expense",
        10,
        account_id="missing-account",
        category_id="missing-category",
    )
    _, workbook = open_export(tmp_path, [transaction])

    assert workbook["Transactions"]["F2"].value is None
    assert workbook["Transactions"]["G2"].value is None
    assert workbook["Category Summary"]["A2"].value is None
    workbook.close()


def test_summary_contains_financial_totals_balance_and_counts(tmp_path) -> None:
    transactions = [
        make_transaction(1, "income", 100.10),
        make_transaction(2, "income", 20),
        make_transaction(3, "expense", 19.99),
    ]
    _, workbook = open_export(tmp_path, transactions)
    summary = {
        row[0].value: row[1].value
        for row in workbook["Summary"].iter_rows(min_row=2)
    }

    assert summary == {
        "Total Income": pytest.approx(120.10),
        "Total Expense": pytest.approx(19.99),
        "Balance": pytest.approx(100.11),
        "Transaction Count": 3,
        "Income Transaction Count": 2,
        "Expense Transaction Count": 1,
    }
    workbook.close()


def test_category_summary_groups_type_separately_and_sorts_names(
    tmp_path,
) -> None:
    transactions = [
        make_transaction(
            1, "expense", 10, category="zeta", category_id=None
        ),
        make_transaction(
            2, "income", 20, category="Alpha", category_id=None
        ),
        make_transaction(
            3, "expense", 5, category="Alpha", category_id=None
        ),
        make_transaction(
            4, "expense", 2.5, category="Alpha", category_id=None
        ),
    ]
    _, workbook = open_export(tmp_path, transactions)
    rows = list(
        workbook["Category Summary"].iter_rows(
            min_row=2,
            values_only=True,
        )
    )

    assert rows == [
        ("Alpha", "Expense", 2, 7.5),
        ("Alpha", "Income", 1, 20),
        ("zeta", "Expense", 1, 10),
    ]
    workbook.close()


def test_empty_export_has_headers_and_zero_summary(tmp_path) -> None:
    _, workbook = open_export(tmp_path, [])

    assert workbook["Transactions"].max_row == 1
    assert workbook["Category Summary"].max_row == 1
    assert [
        row[1].value
        for row in workbook["Summary"].iter_rows(min_row=2)
    ] == [0, 0, 0, 0, 0, 0]
    workbook.close()


def test_export_does_not_mutate_transactions_or_sequence(tmp_path) -> None:
    transactions = [
        make_transaction(2, "expense", 10),
        make_transaction(1, "income", 20),
    ]
    before = deepcopy(transactions)

    export_transactions_to_excel(transactions, tmp_path / "report.xlsx")

    assert transactions == before


def test_existing_destination_is_not_silently_overwritten(tmp_path) -> None:
    destination = tmp_path / "report.xlsx"
    destination.write_bytes(b"original")

    with pytest.raises(ExcelDestinationExistsError):
        export_transactions_to_excel([], destination)

    assert destination.read_bytes() == b"original"


def test_explicit_overwrite_replaces_existing_destination(tmp_path) -> None:
    destination = tmp_path / "report.xlsx"
    destination.write_bytes(b"original")

    export_transactions_to_excel([], destination, overwrite=True)

    workbook = load_workbook(destination)
    assert workbook.sheetnames[0] == "Transactions"
    workbook.close()


def test_failed_save_leaves_no_final_or_temporary_file(
    tmp_path,
    monkeypatch,
) -> None:
    def fail_save(self, filename):
        raise OSError("disk unavailable")

    monkeypatch.setattr(excel_exporter.Workbook, "save", fail_save)
    destination = tmp_path / "report.xlsx"

    with pytest.raises(ExcelSaveError, match="disk unavailable"):
        export_transactions_to_excel([], destination)

    assert not destination.exists()
    assert list(tmp_path.iterdir()) == []


def test_failed_overwrite_preserves_existing_final_file(
    tmp_path,
    monkeypatch,
) -> None:
    destination = tmp_path / "report.xlsx"
    destination.write_bytes(b"original")

    def fail_save(self, filename):
        raise OSError("disk unavailable")

    monkeypatch.setattr(excel_exporter.Workbook, "save", fail_save)
    with pytest.raises(ExcelSaveError):
        export_transactions_to_excel([], destination, overwrite=True)

    assert destination.read_bytes() == b"original"
    assert list(tmp_path.iterdir()) == [destination]


def test_destination_without_extension_is_normalized_and_parents_created(
    tmp_path,
) -> None:
    result = export_transactions_to_excel(
        [],
        tmp_path / "nested" / "report",
    )

    assert result == tmp_path / "nested" / "report.xlsx"
    assert result.is_file()


@pytest.mark.parametrize(
    "destination",
    ["report.csv", "report.xls", "."],
)
def test_invalid_destination_is_rejected(tmp_path, destination) -> None:
    with pytest.raises(InvalidExcelDestinationError):
        export_transactions_to_excel([], tmp_path / destination)
