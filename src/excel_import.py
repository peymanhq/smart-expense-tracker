"""Defensive parsing for the supported Excel transaction workbook."""

from dataclasses import dataclass
from datetime import date, datetime
import math
from pathlib import Path
from zipfile import BadZipFile

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from excel_workbook import (
    LEGACY_EXPORTED_DATE_HEADER,
    REQUIRED_TRANSACTION_HEADERS,
    TRANSACTIONS_WORKSHEET,
)
from validators import (
    validate_amount,
    validate_transaction_date,
    validate_transaction_type,
)


class ExcelImportError(Exception):
    """Base class for controlled Excel import failures."""


class ExcelImportFileNotFoundError(ExcelImportError, FileNotFoundError):
    """Raised when the requested workbook does not exist."""


class UnsupportedExcelImportFileError(ExcelImportError, ValueError):
    """Raised when the source is not an .xlsx file."""


class InvalidExcelWorkbookError(ExcelImportError, ValueError):
    """Raised when an .xlsx file cannot be read as a workbook."""


class MissingTransactionsWorksheetError(ExcelImportError, ValueError):
    """Raised when the required Transactions worksheet is absent."""


class InvalidExcelHeadersError(ExcelImportError, ValueError):
    """Raised when the Transactions worksheet headers are unusable."""


@dataclass(frozen=True)
class ExcelImportIssue:
    """One row-level validation or duplicate conflict."""

    row_number: int
    message: str
    code: str
    field: str | None = None
    supplied_value: object | None = None
    matching_display_id: str | None = None
    earlier_row_number: int | None = None


@dataclass(frozen=True)
class ExcelImportRow:
    """One parsed row before managed-reference resolution."""

    row_number: int
    transaction_date: date
    transaction_type: str
    amount: float
    description: str
    account_name: str
    category_name: str


@dataclass(frozen=True)
class ParsedExcelImport:
    """Workbook parsing result independent from persisted transactions."""

    source_path: Path
    worksheet_name: str
    total_physical_data_rows: int
    empty_rows_ignored: int
    rows: tuple[ExcelImportRow, ...]
    issues: tuple[ExcelImportIssue, ...]


def _header_key(value: object) -> str | None:
    if not isinstance(value, str):
        return None
    key = value.strip().casefold()
    return key or None


def _source_path(source: Path | str) -> Path:
    path = Path(source).expanduser()
    if path.suffix.casefold() != ".xlsx":
        raise UnsupportedExcelImportFileError(
            "Excel import source must use the .xlsx extension."
        )
    if not path.is_file():
        raise ExcelImportFileNotFoundError(
            f"Excel import source was not found: {path}"
        )
    return path


def _header_indexes(header_values: tuple[object, ...]) -> dict[str, int]:
    normalized: dict[str, int] = {}
    duplicates: list[str] = []
    for index, value in enumerate(header_values):
        key = _header_key(value)
        if key is None:
            continue
        if key in normalized:
            duplicates.append(str(value).strip())
        else:
            normalized[key] = index
    if duplicates:
        labels = ", ".join(sorted(set(duplicates), key=str.casefold))
        raise InvalidExcelHeadersError(
            f"Transactions worksheet has duplicate normalized headers: {labels}."
        )

    required_keys = {
        header: header.casefold() for header in REQUIRED_TRANSACTION_HEADERS
    }
    date_key = required_keys["Date"]
    legacy_date_key = LEGACY_EXPORTED_DATE_HEADER.casefold()
    if date_key in normalized and legacy_date_key in normalized:
        raise InvalidExcelHeadersError(
            "Transactions worksheet has ambiguous Date and Transaction Date "
            "headers; keep only the canonical Date column."
        )
    if (
        date_key not in normalized
        and legacy_date_key in normalized
    ):
        normalized[date_key] = normalized[legacy_date_key]

    missing = [
        header
        for header, key in required_keys.items()
        if key not in normalized
    ]
    if missing:
        raise InvalidExcelHeadersError(
            "Transactions worksheet is missing required column(s): "
            + ", ".join(missing)
            + "."
        )
    return {
        header: normalized[key] for header, key in required_keys.items()
    }


def _is_blank(value: object) -> bool:
    return value is None or (isinstance(value, str) and not value.strip())


def _issue(
    row_number: int,
    field: str,
    message: str,
    value: object,
) -> ExcelImportIssue:
    return ExcelImportIssue(
        row_number=row_number,
        field=field,
        message=message,
        code=f"invalid_{field.casefold()}",
        supplied_value=value,
    )


def _required_text(
    value: object,
    row_number: int,
    field: str,
) -> tuple[str | None, ExcelImportIssue | None]:
    if _is_blank(value):
        return None, _issue(
            row_number,
            field,
            f"{field} is required.",
            value,
        )
    if not isinstance(value, str):
        return None, _issue(
            row_number,
            field,
            f"{field} must be text.",
            value,
        )
    return value.strip(), None


def _parse_date(
    value: object,
    row_number: int,
) -> tuple[date | None, ExcelImportIssue | None]:
    if _is_blank(value):
        return None, _issue(
            row_number,
            "Date",
            "Date is required; formula cells need a usable cached value.",
            value,
        )
    if isinstance(value, bool):
        return None, _issue(
            row_number,
            "Date",
            "Date must be a real Excel date or supported YYYY-MM-DD text.",
            value,
        )
    if isinstance(value, datetime):
        return value.date(), None
    try:
        return validate_transaction_date(value), None
    except ValueError:
        return None, _issue(
            row_number,
            "Date",
            "Date must be a valid Excel date or supported YYYY-MM-DD text.",
            value,
        )


def _parse_type(
    value: object,
    row_number: int,
) -> tuple[str | None, ExcelImportIssue | None]:
    if _is_blank(value):
        return None, _issue(
            row_number,
            "Type",
            "Type is required.",
            value,
        )
    if not isinstance(value, str):
        return None, _issue(
            row_number,
            "Type",
            "Type must be Income or Expense.",
            value,
        )
    try:
        return validate_transaction_type(value), None
    except ValueError:
        return None, _issue(
            row_number,
            "Type",
            "Type must be Income or Expense.",
            value,
        )


def _parse_amount(
    value: object,
    row_number: int,
) -> tuple[float | None, ExcelImportIssue | None]:
    if _is_blank(value):
        return None, _issue(
            row_number,
            "Amount",
            "Amount is required; formula cells need a usable cached value.",
            value,
        )
    if isinstance(value, bool):
        return None, _issue(
            row_number,
            "Amount",
            "Amount must be a finite number greater than zero.",
            value,
        )
    try:
        amount = validate_amount(value)
    except (TypeError, ValueError):
        return None, _issue(
            row_number,
            "Amount",
            "Amount must be a valid number greater than zero.",
            value,
        )
    if not math.isfinite(amount):
        return None, _issue(
            row_number,
            "Amount",
            "Amount must be a finite number greater than zero.",
            value,
        )
    return amount, None


def _parse_description(
    value: object,
    row_number: int,
) -> tuple[str | None, ExcelImportIssue | None]:
    if value is None:
        return "", None
    if not isinstance(value, str):
        return None, _issue(
            row_number,
            "Description",
            "Description must be text.",
            value,
        )
    return value.strip(), None


def _parse_row(
    row_number: int,
    values: dict[str, object],
) -> tuple[ExcelImportRow | None, list[ExcelImportIssue]]:
    issues: list[ExcelImportIssue] = []
    transaction_date, issue = _parse_date(values["Date"], row_number)
    if issue is not None:
        issues.append(issue)
    transaction_type, issue = _parse_type(values["Type"], row_number)
    if issue is not None:
        issues.append(issue)
    amount, issue = _parse_amount(values["Amount"], row_number)
    if issue is not None:
        issues.append(issue)
    description, issue = _parse_description(
        values["Description"],
        row_number,
    )
    if issue is not None:
        issues.append(issue)
    account_name, issue = _required_text(
        values["Account"],
        row_number,
        "Account",
    )
    if issue is not None:
        issues.append(issue)
    category_name, issue = _required_text(
        values["Category"],
        row_number,
        "Category",
    )
    if issue is not None:
        issues.append(issue)

    if issues:
        return None, issues
    assert transaction_date is not None
    assert transaction_type is not None
    assert amount is not None
    assert description is not None
    assert account_name is not None
    assert category_name is not None
    return (
        ExcelImportRow(
            row_number=row_number,
            transaction_date=transaction_date,
            transaction_type=transaction_type,
            amount=amount,
            description=description,
            account_name=account_name,
            category_name=category_name,
        ),
        [],
    )


def parse_excel_transactions(source: Path | str) -> ParsedExcelImport:
    """Parse all supported rows while retaining their physical row numbers."""
    source_path = _source_path(source)
    workbook = None
    try:
        workbook = load_workbook(
            source_path,
            read_only=True,
            data_only=True,
            keep_links=False,
        )
    except (InvalidFileException, BadZipFile, KeyError, OSError, ValueError) as error:
        raise InvalidExcelWorkbookError(
            f"Could not read Excel workbook {source_path}: {error}"
        ) from error

    try:
        if TRANSACTIONS_WORKSHEET not in workbook.sheetnames:
            raise MissingTransactionsWorksheetError(
                "Excel workbook must contain a worksheet named exactly "
                f"{TRANSACTIONS_WORKSHEET}."
            )
        worksheet = workbook[TRANSACTIONS_WORKSHEET]
        header_values = tuple(
            cell.value for cell in next(worksheet.iter_rows(min_row=1, max_row=1))
        )
        indexes = _header_indexes(header_values)

        rows: list[ExcelImportRow] = []
        issues: list[ExcelImportIssue] = []
        empty_rows_ignored = 0
        total_physical_data_rows = max(worksheet.max_row - 1, 0)
        for row_number, cells in enumerate(
            worksheet.iter_rows(min_row=2),
            start=2,
        ):
            values = {
                header: (
                    cells[index].value if index < len(cells) else None
                )
                for header, index in indexes.items()
            }
            if all(_is_blank(value) for value in values.values()):
                empty_rows_ignored += 1
                continue
            parsed_row, row_issues = _parse_row(row_number, values)
            issues.extend(row_issues)
            if parsed_row is not None:
                rows.append(parsed_row)

        return ParsedExcelImport(
            source_path=source_path,
            worksheet_name=TRANSACTIONS_WORKSHEET,
            total_physical_data_rows=total_physical_data_rows,
            empty_rows_ignored=empty_rows_ignored,
            rows=tuple(rows),
            issues=tuple(issues),
        )
    finally:
        workbook.close()
