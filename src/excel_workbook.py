"""Shared Excel workbook contracts and safe output helpers."""

from collections.abc import Sequence
import os
from pathlib import Path
import tempfile

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

TRANSACTIONS_WORKSHEET = "Transactions"
REQUIRED_TRANSACTION_HEADERS = (
    "Date",
    "Type",
    "Amount",
    "Description",
    "Account",
    "Category",
)
EXPORTED_TRANSACTION_HEADERS = (
    "Display ID",
    *REQUIRED_TRANSACTION_HEADERS,
    "Created At",
    "Updated At",
)
LEGACY_EXPORTED_DATE_HEADER = "Transaction Date"

AMOUNT_FORMAT = "#,##0.00"
DATE_FORMAT = "yyyy-mm-dd"
DATETIME_FORMAT = "yyyy-mm-dd hh:mm:ss"
HEADER_FILL_COLOR = "1F4E78"
HEADER_FONT_COLOR = "FFFFFF"


class ExcelWorkbookError(Exception):
    """Base class for controlled Excel workbook output failures."""


class InvalidExcelDestinationError(ExcelWorkbookError, ValueError):
    """Raised when a destination cannot represent an Excel workbook."""


class ExcelDestinationExistsError(ExcelWorkbookError, FileExistsError):
    """Raised when output would overwrite a file without permission."""


class ExcelSaveError(ExcelWorkbookError, OSError):
    """Raised when a workbook cannot be written safely."""


def normalize_excel_destination(destination: Path | str) -> Path:
    """Return a normalized .xlsx path or reject a misleading extension."""
    path = Path(destination).expanduser()
    if not path.name or path.name in {".", ".."}:
        raise InvalidExcelDestinationError("Excel destination must name a file.")
    if path.exists() and path.is_dir():
        raise InvalidExcelDestinationError(
            "Excel destination must name a file, not a directory."
        )
    if not path.suffix:
        path = path.with_suffix(".xlsx")
    elif path.suffix.casefold() != ".xlsx":
        raise InvalidExcelDestinationError(
            "Excel destination must use the .xlsx extension."
        )
    return path


def style_header_row(worksheet: Worksheet) -> None:
    """Apply the established workbook header style to row one."""
    header_fill = PatternFill("solid", fgColor=HEADER_FILL_COLOR)
    header_font = Font(color=HEADER_FONT_COLOR, bold=True)
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")


def style_table(
    worksheet: Worksheet,
    widths: Sequence[float],
    *,
    wrap_columns: set[int] | None = None,
) -> None:
    """Apply the shared table header, navigation, and width conventions."""
    style_header_row(worksheet)
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    for index, width in enumerate(widths, start=1):
        worksheet.column_dimensions[
            worksheet.cell(row=1, column=index).column_letter
        ].width = width
    for column in wrap_columns or set():
        for row in range(2, worksheet.max_row + 1):
            worksheet.cell(row=row, column=column).alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )


def save_workbook_atomic(
    workbook: Workbook,
    destination: Path | str,
    *,
    overwrite: bool = False,
) -> Path:
    """Save a workbook through one same-directory atomic replacement."""
    destination_path: Path | None = None
    temporary_path: Path | None = None
    try:
        destination_path = normalize_excel_destination(destination)
        if destination_path.exists() and not overwrite:
            raise ExcelDestinationExistsError(
                f"Excel destination already exists: {destination_path}"
            )
        destination_path.parent.mkdir(parents=True, exist_ok=True)
        with tempfile.NamedTemporaryFile(
            prefix=f".{destination_path.stem}.",
            suffix=".xlsx",
            dir=destination_path.parent,
            delete=False,
        ) as temporary_file:
            temporary_path = Path(temporary_file.name)
        workbook.save(temporary_path)
        os.replace(temporary_path, destination_path)
    except ExcelWorkbookError:
        raise
    except (OSError, ValueError) as error:
        raise ExcelSaveError(
            f"Could not save Excel workbook to "
            f"{destination_path or destination}: {error}"
        ) from error
    finally:
        workbook.close()
        if temporary_path is not None:
            temporary_path.unlink(missing_ok=True)
    assert destination_path is not None
    return destination_path
