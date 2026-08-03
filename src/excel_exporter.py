"""Excel reporting adapter for in-memory transaction collections."""

from collections import defaultdict
from collections.abc import Mapping, Sequence
from datetime import datetime, timezone
from decimal import Decimal
from pathlib import Path
import unicodedata

from openpyxl import Workbook

from excel_workbook import (
    AMOUNT_FORMAT,
    DATE_FORMAT,
    DATETIME_FORMAT,
    EXPORTED_TRANSACTION_HEADERS,
    ExcelDestinationExistsError,
    ExcelSaveError,
    ExcelWorkbookError,
    InvalidExcelDestinationError,
    normalize_excel_destination,
    save_workbook_atomic,
    style_table,
)
from id_generator import parse_display_id
from report import calculate_financial_summary
from transaction import Transaction

TRANSACTION_HEADERS = EXPORTED_TRANSACTION_HEADERS
SUMMARY_HEADERS = ("Metric", "Value")
CATEGORY_SUMMARY_HEADERS = (
    "Category",
    "Type",
    "Transaction Count",
    "Total Amount",
)
ExcelExportError = ExcelWorkbookError


def _transaction_order(transaction: Transaction) -> tuple:
    display_number = parse_display_id(transaction.display_id)
    return (
        transaction.transaction_date,
        display_number if display_number is not None else 2**63 - 1,
        transaction.display_id,
    )


def _resolved_name(
    snapshot: str,
    reference_id: str | None,
    names: Mapping[str, str],
) -> str | None:
    if reference_id is None:
        return snapshot
    return names.get(reference_id)


def _excel_datetime(value: datetime | None) -> datetime | None:
    """Convert aware timestamps to Excel-compatible naive UTC datetimes."""
    if value is None or value.tzinfo is None:
        return value
    return value.astimezone(timezone.utc).replace(tzinfo=None)


def _write_transactions(
    workbook: Workbook,
    transactions: Sequence[Transaction],
    account_names: Mapping[str, str],
    category_names: Mapping[str, str],
) -> None:
    worksheet = workbook.create_sheet("Transactions")
    worksheet.append(TRANSACTION_HEADERS)
    for transaction in sorted(transactions, key=_transaction_order):
        worksheet.append(
            (
                transaction.display_id,
                transaction.transaction_date,
                transaction.type.title(),
                float(transaction.amount),
                transaction.description,
                _resolved_name(
                    transaction.account,
                    transaction.account_id,
                    account_names,
                ),
                _resolved_name(
                    transaction.category,
                    transaction.category_id,
                    category_names,
                ),
                _excel_datetime(transaction.created_at),
                _excel_datetime(transaction.updated_at),
            )
        )

    for cell in worksheet["B"][1:]:
        cell.number_format = DATE_FORMAT
    for cell in worksheet["D"][1:]:
        cell.number_format = AMOUNT_FORMAT
    for column in ("H", "I"):
        for cell in worksheet[column][1:]:
            cell.number_format = DATETIME_FORMAT
    style_table(
        worksheet,
        (14, 18, 12, 15, 42, 24, 24, 22, 22),
        wrap_columns={5},
    )


def _write_summary(
    workbook: Workbook,
    transactions: Sequence[Transaction],
) -> None:
    worksheet = workbook.create_sheet("Summary")
    worksheet.append(SUMMARY_HEADERS)
    summary = calculate_financial_summary(transactions)
    income_count = sum(
        transaction.type == "income" for transaction in transactions
    )
    expense_count = sum(
        transaction.type == "expense" for transaction in transactions
    )
    rows = (
        ("Total Income", float(summary.total_income)),
        ("Total Expense", float(summary.total_expense)),
        ("Balance", float(summary.balance)),
        ("Transaction Count", summary.transaction_count),
        ("Income Transaction Count", income_count),
        ("Expense Transaction Count", expense_count),
    )
    for row in rows:
        worksheet.append(row)
    for cell in worksheet["B"][1:4]:
        cell.number_format = AMOUNT_FORMAT
    style_table(worksheet, (30, 20))


def _write_category_summary(
    workbook: Workbook,
    transactions: Sequence[Transaction],
    category_names: Mapping[str, str],
) -> None:
    worksheet = workbook.create_sheet("Category Summary")
    worksheet.append(CATEGORY_SUMMARY_HEADERS)
    groups: dict[tuple[str, str], list[int | Decimal]] = defaultdict(
        lambda: [0, Decimal("0")]
    )
    for transaction in transactions:
        category = _resolved_name(
            transaction.category,
            transaction.category_id,
            category_names,
        )
        category_label = "" if category is None else category
        group = groups[(category_label, transaction.type)]
        group[0] += 1
        group[1] += transaction.amount

    def category_order(item: tuple[str, str]) -> tuple[str, str, str]:
        category, transaction_type = item
        normalized = unicodedata.normalize("NFC", category).casefold()
        return normalized, transaction_type, category

    for category, transaction_type in sorted(groups, key=category_order):
        count, total = groups[(category, transaction_type)]
        worksheet.append(
            (category or None, transaction_type.title(), count, float(total))
        )
    for cell in worksheet["D"][1:]:
        cell.number_format = AMOUNT_FORMAT
    style_table(worksheet, (26, 14, 20, 18))


def _build_workbook(
    transactions: Sequence[Transaction],
    account_names: Mapping[str, str],
    category_names: Mapping[str, str],
) -> Workbook:
    workbook = Workbook()
    workbook.remove(workbook.active)
    _write_transactions(workbook, transactions, account_names, category_names)
    _write_summary(workbook, transactions)
    _write_category_summary(workbook, transactions, category_names)
    return workbook


def export_transactions_to_excel(
    transactions: Sequence[Transaction],
    destination: Path | str,
    *,
    account_names: Mapping[str, str] | None = None,
    category_names: Mapping[str, str] | None = None,
    overwrite: bool = False,
) -> Path:
    """Create an atomic reporting workbook without reading persistence."""
    transaction_snapshot = tuple(transactions)
    workbook = _build_workbook(
        transaction_snapshot,
        {} if account_names is None else account_names,
        {} if category_names is None else category_names,
    )
    return save_workbook_atomic(
        workbook,
        destination,
        overwrite=overwrite,
    )
